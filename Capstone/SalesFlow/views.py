from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.db import IntegrityError
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.shortcuts import render
from django.urls import reverse
from openpyxl import Workbook
from openpyxl.styles import *

import json
from .models import User, MonthSheet, YearSheet

import datetime

@login_required(login_url="login")
def start(request):
    user= request.user
    today = datetime.datetime.now()
    current_year = today.year  
    yearsheets = user.user_yearsheets.all()
    
    # Check if the user has year sheets and if the current year exists in the user year sheets
    # Display the current year if the user has it or if the user has no year sheets, in any other case display any og the year sheets that exists
    if yearsheets:
        for yearsheet in yearsheets:
            if int(yearsheet.year) == int(current_year):
                req_year = current_year
                break
        req_year = yearsheets[0].year    
    else:
        req_year = current_year    
           
    return HttpResponseRedirect(reverse("index", kwargs={'req_year': req_year}))

@login_required(login_url="login")
def index(request, req_year):
 
    # Retrieve current user and the yearsheets of the user.
    user_object = User.objects.get(username=request.user)
    years = user_object.user_yearsheets.all()
    year_list = []
    for year in years:
        year_list.append(year.year)
    
    # Get current calendar year and try to retrieve yearsheet and monthsheets for the year if they exist:
    today = datetime.datetime.now()
    current_year = today.year
    existing_years = True
    
    if req_year in year_list:
        display_year = req_year
        current_yearsheet = user_object.user_yearsheets.get(year=int(req_year))
        months = current_yearsheet.months.all()
    elif current_year in year_list:
        display_year = current_year
        current_yearsheet = user_object.user_yearsheets.get(year=int(current_year))
        months = current_yearsheet.months.all()
    else:
        display_year = current_year
        months = ""
        existing_years = False
        
    return render(request,"SalesFlow/index.html", {
        "username": request.user,
        "year_list": year_list,
        "current_year": display_year,
        "months":months,
        "existing_years": existing_years
    })
    
def edit_month_data(request):
        
    if request.method == 'POST':
        data = json.loads(request.body)
        new_data = data.get("new_data")
        month_id = data.get("month_id")
        field = data.get("field")

        month_sheet = MonthSheet.objects.get(pk=month_id)
        setattr(month_sheet, field, new_data) 
        month_sheet.save()
                
        return JsonResponse({'response': 'data received succesfully', 'data': data}, status=200)

def delete_year(request, year):
    user = request.user
    year_to_delete = user.user_yearsheets.get(year=year)
    year_to_delete.delete() 
    return HttpResponseRedirect(reverse('start'))

def login_view(request):
    today = datetime.datetime.now()
    current_year = today.year

    if request.method == "POST":
        # Attempt to sign user in
        email = request.POST["email"]
        password = request.POST["password"]
        
        try:
            user = User.objects.get(email=email)
            if user.check_password(password):
                username = user.username
                user = authenticate(request, username=username, password=password)
        except User.DoesNotExist:
            return render(request, "SalesFlow/login.html", {
                "message": "verifique los datos ingresados.",
                "current_year": current_year
            })
        
        # Check if authentication was successful
        if user is not None:
            login(request, user)
            return HttpResponseRedirect(reverse("start"))
        else:
            return render(request, "SalesFlow/login.html", {
                "message": "verifique los datos ingresados.",
                "current_year": current_year
            })
    else:
        return render(request, "SalesFlow/login.html", {
            "current_year":current_year
        })

def logout_view(request):
    logout(request)
    today = datetime.datetime.now()
    current_year = today.year
    return HttpResponseRedirect(reverse("index", kwargs={'req_year': current_year}))

def register(request):

    today = datetime.datetime.now()
    current_year = today.year

    if request.method == "POST":
        username = request.POST["full_name"]
        email = request.POST["email"]

        # Ensure password matches confirmation
        password = request.POST["password"]
        confirmation = request.POST["confirmation"]
        if password != confirmation:
            return render(request, "SalesFlow/register.html", {
                "message": "Las contraseñas no coinciden.",
                "current_year": current_year
            })

        # Attempt to create new user
        try:
            user = User.objects.create_user(username, email, password)
            user.save()
           
        except IntegrityError:
            return render(request, "SalesFlow/register.html", {
                "message": "El usuario ya existe, inicie sesion.",
                "current_year": current_year
            })
        
        except ValueError:
            return render(request, "SalesFlow/register.html", {
                "message": "Complete todos los campos.",
                "current_year": current_year
            })
        
        login(request, user)

        return HttpResponseRedirect(reverse("index", kwargs={'req_year': current_year}))
    else:
        return render(request, "SalesFlow/register.html",{
            "current_year": current_year
        })
 
@login_required(login_url="login")    
def new_year(request):

    user = request.user
    today = datetime.datetime.now()
    current_year = today.year
    existing_years = user.user_yearsheets.all()
    year_list = []
    for existing_year in existing_years:
        year_list.append(int(existing_year.year))

    if request.method == 'POST':
        # Load data from API request:
        data = json.loads(request.body)

        # Save data fields in variables:
        year = int(data.get('year'))
        opt_goal = int(data.get('opt_goal'))
        sales_goal = int(data.get('sales_goal'))
        fact_goal = int(data.get('fact_goal'))
        efect_goal = float(data.get('efect_goal'))
        td_goal = int(data.get('td_goal'))
        
        #check if this year already exists:    
        if year in year_list:
            return render(request, "SalesFlow/new_year.html", {
                "message": f"El año {year} ya existe en su cuenta",
                "username": user,
                "current_year": current_year
            })

        # Create Year object in database:
        new_year = YearSheet(
            year = year,
            user = user
        )
        new_year.save()
        year_list.append(int(year))

        # Create Month objects in database:
        months = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']

        for current_month in months:
            new_month = MonthSheet(
                year_sheet = new_year,
                month = current_month,
                opt_goal = opt_goal,
                opt_current = 0,
                sales_goal = sales_goal,
                sales_current = 0,
                fact_goal = fact_goal,
                fact_current = 0,
                efect_goal = efect_goal,
                efect_current = 0,
                td_goal = td_goal,
                td_current = 0
            )
            new_month.save()

        if year in year_list:
            current_yearsheet = user.user_yearsheets.get(year=int(year))
            months = current_yearsheet.months.all()
        else:
            months = ""

        return JsonResponse({'response': 'data received succesfully', 'data': data}, status=200)
        

    return render(request, "SalesFlow/new_year.html", {
        "username": user,
        "current_year": current_year,
        "year_list": year_list
    })

def get_monthsheet(request, pk):
    
    if request.method == 'GET':
        monthsheet = MonthSheet.objects.get(pk=pk)
        monthsheet_data = {
            'month_name': monthsheet.month,
            'opt_current': monthsheet.opt_current,
            'fact_current': monthsheet.fact_current
        }
        return JsonResponse({'month_data': monthsheet_data})
    
def excel(request):
    
    if request.method == 'POST':
        
        # Get the user, year and corresponding month sheets:
        user = request.user
        year = int(request.POST['excel-year'])
        year_sheet = user.user_yearsheets.get(year=year)
        months = year_sheet.months.all()
        yearsheet_data = []
        
        # Create dicts to store each month sheet data:
        for month in months:
            data = {}
            data['month'] = month.month
            data['opt_goal'] = month.opt_goal
            data['opt_current'] = month.opt_current
            data['sales_goal'] = month.sales_goal
            data['sales_current'] = month.sales_current
            data['fact_goal'] = month.fact_goal
            data['fact_current'] = month.fact_current
            data['efect_goal'] = month.efect_goal
            data['efect_current'] = month.efect_current
            data['td_goal'] = month.td_goal
            data['td_current'] = month.td_current
            yearsheet_data.append(data)

        # Create excel book:
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',)
        response['Content-Disposition'] = 'attachment; filename="' + 'Cuadro de seguimiento - ' + str(year) + ' ' + user.username +'.xlsx"'
        workbook = Workbook()
        worksheet = workbook.active
        
        # Fill the headers cells:
        name_cell = worksheet['B6']
        name_cell.value = user.username
        worksheet.merge_cells('B6:Q6')
        name_cell.font = Font(bold=True, size=12)
        name_cell.alignment = Alignment(horizontal="center", vertical="center")
                
        year_cell = worksheet['B7']
        year_cell.value = f'AÑO {year}'
        year_cell.font = Font(bold=True)
        year_cell.alignment = Alignment(horizontal="center", vertical="center")
               
        first_headers = ['OPORT. SEM', 'VENTAS SEM', 'FACTURAS SEM', '% EFECT', 'P. MANEJO']
        sub_headers = ['EJECT', 'META', '% CUMP']
        col = 3
        row = 7
        
        for header in first_headers:
            cell = worksheet.cell(row, col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            for sub_header in sub_headers:
                cell = worksheet.cell(row + 1, col)
                cell.value = sub_header
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                col += 1
        
        worksheet.merge_cells('B7:B8') 
        worksheet.merge_cells('C7:E7')
        worksheet.merge_cells('F7:H7')
        worksheet.merge_cells('I7:K7')
        worksheet.merge_cells('L7:N7')
        worksheet.merge_cells('O7:Q7')

        # Fill the table with all the data:
        row = 9
        quarter = 1
        for month in yearsheet_data:
            if row in [12,16,20,24]:
                worksheet.cell(row, 2).value = 'Q' + str(quarter)
                worksheet.cell(row, 3).value = worksheet.cell(row - 1, 3).value + worksheet.cell(row - 2, 3).value + worksheet.cell(row - 3, 3).value
                worksheet.cell(row, 4).value = worksheet.cell(row - 1, 4).value + worksheet.cell(row - 2, 4).value + worksheet.cell(row - 3, 4).value
                worksheet.cell(row, 5).value = worksheet.cell(row, 3).value / worksheet.cell(row, 4).value if worksheet.cell(row, 4).value > 0 else 0 
                worksheet.cell(row, 6).value = worksheet.cell(row - 1, 6).value + worksheet.cell(row - 2, 6).value + worksheet.cell(row - 3, 6).value
                worksheet.cell(row, 7).value = worksheet.cell(row - 1, 7).value + worksheet.cell(row - 2, 7).value + worksheet.cell(row - 3, 7).value
                worksheet.cell(row, 8).value = worksheet.cell(row, 6).value / worksheet.cell(row, 7).value if worksheet.cell(row, 7).value > 0 else 0
                worksheet.cell(row, 9).value = worksheet.cell(row - 1, 9).value + worksheet.cell(row - 2, 9).value + worksheet.cell(row - 3, 9).value
                worksheet.cell(row, 10).value = worksheet.cell(row - 1, 10).value + worksheet.cell(row - 2, 10).value + worksheet.cell(row - 3, 10).value
                worksheet.cell(row, 11).value = worksheet.cell(row, 9).value / worksheet.cell(row, 10).value if worksheet.cell(row, 10).value > 0 else 0
                worksheet.cell(row, 12).value = worksheet.cell(row, 9).value / worksheet.cell(row, 3).value if worksheet.cell(row, 3).value > 0 else 0
                worksheet.cell(row, 13).value = worksheet.cell(row - 1, 13).value
                worksheet.cell(row, 14).value = worksheet.cell(row, 13).value / worksheet.cell(row, 12).value if worksheet.cell(row, 12).value > 0 else 0 
                worksheet.cell(row, 15).value = worksheet.cell(row - 1, 15).value + worksheet.cell(row - 2, 15).value + worksheet.cell(row - 3, 15).value
                worksheet.cell(row, 16).value = worksheet.cell(row - 1, 16).value + worksheet.cell(row - 2, 16).value + worksheet.cell(row - 3, 16).value
                worksheet.cell(row, 17).value = worksheet.cell(row, 15).value / worksheet.cell(row, 16).value if worksheet.cell(row, 16).value > 0 else 0 

                quarter += 1

                worksheet.cell(row + 1, 2).value = month['month']
                worksheet.cell(row + 1, 3).value = month['opt_current']
                worksheet.cell(row + 1, 4).value = month['opt_goal']
                worksheet.cell(row + 1, 5).value = month['opt_current']/month['opt_goal'] if month['opt_goal'] > 0 else 0
                worksheet.cell(row + 1, 6).value = month['sales_current']
                worksheet.cell(row + 1, 7).value = month['sales_goal']
                worksheet.cell(row + 1, 8).value = month['sales_current']/month['sales_goal'] if month['sales_goal'] > 0 else 0
                worksheet.cell(row + 1, 9).value = month['fact_current']
                worksheet.cell(row + 1, 10).value = month['fact_goal']
                worksheet.cell(row + 1, 11).value = month['fact_current']/month['fact_goal'] if month['fact_goal'] > 0 else 0
                worksheet.cell(row + 1, 12).value = month['efect_current'] / 100
                worksheet.cell(row + 1, 13).value = month['efect_goal'] / 100
                worksheet.cell(row + 1, 14).value = month['efect_goal']/month['efect_current'] if month['efect_current'] > 0 else 0
                worksheet.cell(row + 1, 15).value = month['td_current']
                worksheet.cell(row + 1, 16).value = month['td_goal']
                worksheet.cell(row + 1, 17).value = month['td_current']/month['td_goal'] if month['td_goal'] > 0 else 0
                row += 2
            else:
                worksheet.cell(row, 2).value = month['month']
                worksheet.cell(row, 3).value = month['opt_current']
                worksheet.cell(row, 4).value = month['opt_goal']
                worksheet.cell(row, 5).value = month['opt_current']/month['opt_goal'] if month['opt_goal'] > 0 else 0
                worksheet.cell(row, 6).value = month['sales_current']
                worksheet.cell(row, 7).value = month['sales_goal']
                worksheet.cell(row, 8).value = month['sales_current']/month['sales_goal'] if month['sales_goal'] > 0 else 0
                worksheet.cell(row, 9).value = month['fact_current']
                worksheet.cell(row, 10).value = month['fact_goal']
                worksheet.cell(row, 11).value = month['fact_current']/month['fact_goal'] if month['fact_goal'] > 0 else 0
                worksheet.cell(row, 12).value = month['efect_current'] / 100
                worksheet.cell(row, 13).value = month['efect_goal'] / 100
                worksheet.cell(row, 14).value = month['efect_goal']/month['efect_current'] if month['efect_current'] > 0 else 0
                worksheet.cell(row ,15).value = month['td_current']
                worksheet.cell(row, 16).value = month['td_goal']
                worksheet.cell(row, 17).value = month['td_current']/month['td_goal'] if month['td_goal'] > 0 else 0
                row += 1 

        # Fill data for Q4 an  TOTALS:
        worksheet.cell(row, 2).value = 'Q' + str(quarter)
        worksheet.cell(row, 3).value = worksheet.cell(row - 1, 3).value + worksheet.cell(row - 2, 3).value + worksheet.cell(row - 3, 3).value
        worksheet.cell(row, 4).value = worksheet.cell(row - 1, 4).value + worksheet.cell(row - 2, 4).value + worksheet.cell(row - 3, 4).value
        worksheet.cell(row, 5).value = worksheet.cell(row, 3).value / worksheet.cell(row, 4).value if worksheet.cell(row, 4).value > 0 else 0 
        worksheet.cell(row, 6).value = worksheet.cell(row - 1, 6).value + worksheet.cell(row - 2, 6).value + worksheet.cell(row - 3, 6).value
        worksheet.cell(row, 7).value = worksheet.cell(row - 1, 7).value + worksheet.cell(row - 2, 7).value + worksheet.cell(row - 3, 7).value
        worksheet.cell(row, 8).value = worksheet.cell(row, 6).value / worksheet.cell(row, 7).value if worksheet.cell(row, 7).value > 0 else 0
        worksheet.cell(row, 9).value = worksheet.cell(row - 1, 9).value + worksheet.cell(row - 2, 9).value + worksheet.cell(row - 3, 9).value
        worksheet.cell(row, 10).value = worksheet.cell(row - 1, 10).value + worksheet.cell(row - 2, 10).value + worksheet.cell(row - 3, 10).value
        worksheet.cell(row, 11).value = worksheet.cell(row, 9).value / worksheet.cell(row, 10).value if worksheet.cell(row, 10).value > 0 else 0
        worksheet.cell(row, 12).value = worksheet.cell(row, 9).value / worksheet.cell(row, 3).value if worksheet.cell(row, 3).value > 0 else 0
        worksheet.cell(row, 13).value = worksheet.cell(row - 1, 13).value
        worksheet.cell(row, 14).value = worksheet.cell(row, 13).value / worksheet.cell(row, 12).value if worksheet.cell(row, 12).value > 0 else 0 
        worksheet.cell(row, 15).value = worksheet.cell(row - 1, 15).value + worksheet.cell(row - 2, 15).value + worksheet.cell(row - 3, 15).value
        worksheet.cell(row, 16).value = worksheet.cell(row - 1, 16).value + worksheet.cell(row - 2, 16).value + worksheet.cell(row - 3, 16).value
        worksheet.cell(row, 17).value = worksheet.cell(row, 15).value / worksheet.cell(row, 16).value if worksheet.cell(row, 16).value > 0 else 0 


        worksheet.cell(25, 2).value = 'TOTAL:'
        worksheet.cell(25, 3).value = '=C24 + C20 + C16 + C12'
        worksheet.cell(25, 4).value = '=D24 + D20 + D16 + D12'
        worksheet.cell(25, 5).value = '=C25 / D25'
        worksheet.cell(25, 6).value = '=F24 + F20 + F16 + F12'
        worksheet.cell(25, 7).value = '=G24 + G20 + G16 + G12'
        worksheet.cell(25, 8).value = '=F25 / G25'
        worksheet.cell(25, 9).value = '=I24 + I20 + I16 + I12'
        worksheet.cell(25, 10).value ='=J24 + J20 + J16 + J12' 
        worksheet.cell(25, 11).value = '=I5 / J25'
        worksheet.cell(25, 12).value = '=I25 / C25'
        worksheet.cell(25, 13).value = '=M9'
        worksheet.cell(25,14).value = '=M25 / L25'
        worksheet.cell(25, 15).value = '=O24 + O20 + O16 + O12'
        worksheet.cell(25, 16).value = '=P24 + P20 + P16 + P12'
        worksheet.cell(25, 17).value = '=O25 / P25'

        
        # format cells:
        border_object = Border(top=Side(border_style="medium", color="FF000000"),
                           bottom=Side(border_style="medium", color="FF000000"),
                           left=Side(border_style="medium", color="FF000000"),
                           right=Side(border_style="medium", color="FF000000"))

        for row in worksheet['B6:Q25']:
            for cell in row:
                cell.border = border_object
                cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in worksheet['B9:B25']:
            for cell in row:
                cell.alignment = Alignment(horizontal="left", vertical="center")

        for row_num in ['B12:Q12', 'B16:Q16', 'B20:Q20', 'B24:Q24']:
            for row in worksheet[row_num]:
                for cell in row:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill("solid", fgColor="D9D9D9")
                
        for row in worksheet['B25:Q25']:
            for cell in row:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="FFFFFF")
            
        for range in ['E9:E25','H9:H25','K9:K25','L9:L25', 'M9:M25','N9:N25', 'Q9:Q25']:
            for row in worksheet[range]:
                for cell in row:
                    cell.number_format = '0.00%'
        
        for range in ['C9:C24', 'F9:F24', 'I9:I24', 'L9:L24', 'O9:O24']:
            for row in worksheet[range]:
                for cell in row:
                    cell.fill = PatternFill("solid", fgColor="D9D9D9")

        worksheet.column_dimensions['A'].width = 3
        worksheet.column_dimensions['B'].width = 15
        worksheet.row_dimensions[6].height = 40


        workbook.save(response)      
        return response
    
    return HttpResponse('excel')